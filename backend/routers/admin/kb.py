from fastapi import APIRouter, Depends, HTTPException, File, UploadFile
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import List
import os
import shutil
import csv
from datetime import datetime

from ...database import get_db
from ...models import User
from ...auth.dependencies import require_role
from ...ingest_knowledge_base import ingest_directory

# Resolve backend/knowledge_base path dynamically from backend/routers/admin/kb.py
BACKEND_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
KB_DIR = os.path.join(BACKEND_DIR, "knowledge_base")

# Ensure directory exists
if not os.path.exists(KB_DIR):
    os.makedirs(KB_DIR)

router = APIRouter()

@router.post("/reindex")
def trigger_kb_reindex(
    admin: User = Depends(require_role("admin"))
):
    """Manually trigger re-indexing of the knowledge base directory"""
    kb_dir = KB_DIR
    try:
        ingest_directory(kb_dir)
        return {"message": "Knowledge base re-indexing triggered successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Re-indexing failed: {str(e)}")


@router.get("/kb/files")
def list_kb_files(
    admin: User = Depends(require_role("admin"))
):
    """List all files in the knowledge base directory"""
    kb_dir = KB_DIR
    files = []
    for filename in os.listdir(kb_dir):
        file_path = os.path.join(kb_dir, filename)
        if os.path.isfile(file_path):
            stats = os.stat(file_path)
            files.append({
                "name": filename,
                "size": stats.st_size,
                "modified": datetime.fromtimestamp(stats.st_mtime)
            })
    return files


@router.post("/kb/upload")
async def upload_kb_files(
    files: List[UploadFile] = File(...),
    admin: User = Depends(require_role("admin"))
):
    """Upload one or more files to the knowledge base"""
    kb_dir = KB_DIR
    uploaded_files = []
    for file in files:
        # Secure filename or at least check extension
        if not file.filename.endswith(('.xlsx', '.csv', '.py')):
            continue
            
        file_path = os.path.join(kb_dir, file.filename)
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        uploaded_files.append(file.filename)
        
    return {"message": f"Uploaded {len(uploaded_files)} files: {', '.join(uploaded_files)}"}


@router.get("/kb/files/{filename}/preview")
def preview_kb_file(
    filename: str,
    admin: User = Depends(require_role("employee"))
):
    """Return file contents as JSON rows for in-browser preview (CSV and XLSX only)."""
    # Sanitize — no path traversal
    if "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")

    file_path = os.path.join(KB_DIR, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")

    ext = filename.lower().rsplit(".", 1)[-1]

    try:
        if ext == "csv":
            rows = []
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames or []
                for row in reader:
                    rows.append(dict(row))
            return {"filename": filename, "type": "csv", "headers": list(headers), "rows": rows[:500]}

        elif ext == "xlsx":
            try:
                import openpyxl
            except ImportError:
                raise HTTPException(status_code=500, detail="openpyxl not installed")
            
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheets = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_data = []
                headers = []
                # Determine max columns for headers
                
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        # Normalize headers
                        headers = [str(c) if c is not None and str(c).strip() != "" else f"Column {idx+1}" for idx, c in enumerate(row)]
                    else:
                        # Safely map values to available headers
                        row_dict = {}
                        for j, value in enumerate(row):
                            if j < len(headers):
                                row_dict[headers[j]] = str(value) if value is not None else ""
                            else:
                                pass
                        rows_data.append(row_dict)
                    
                    if i >= 500:
                        break
                
                sheets[sheet_name] = {"headers": headers, "rows": rows_data}
            wb.close()
            return {"filename": filename, "type": "xlsx", "sheets": sheets}

        else:
            raise HTTPException(status_code=400, detail="Only CSV and XLSX preview supported")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Preview failed: {str(e)}")


@router.get("/kb/files/{filename}/download")
def download_kb_file(
    filename: str,
    admin: User = Depends(require_role("employee"))
):
    """Download a KB file directly."""
    if "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    file_path = os.path.join(KB_DIR, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(
        path=file_path,
        filename=filename,
        media_type="application/octet-stream"
    )


@router.delete("/kb/files/{filename}")
def delete_kb_file(
    filename: str,
    admin: User = Depends(require_role("admin"))
):
    """Delete a file from the knowledge base"""
    file_path = os.path.join(KB_DIR, filename)
    
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")
        
    try:
        os.remove(file_path)
        return {"message": f"File {filename} deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to delete file: {str(e)}")
