from fastapi import APIRouter, Depends, HTTPException, status, File, UploadFile
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, cast, String
from typing import List, Dict
import os
import psutil
import shutil
import csv
import json
from datetime import datetime

from ..database import get_db
from ..models import User, UserProgress, QuizScore, SystemLog, Broadcast, ChatLog, ChatFeedback, QueryCache, Quiz, Question, Course, Lesson, LessonContent, QuestionAttempt
from ..schemas import (
    UserCreateAdmin, UserUpdate, UserResponse, 
    QuizCreate, QuizUpdate, QuizResponse, 
    QuestionCreate, QuestionUpdate, QuestionResponse,
    CourseCreate, CourseResponse, 
    LessonCreate, LessonResponse, 
    LessonContentCreate, LessonContentResponse
)
from ..auth.dependencies import require_role
from ..auth.security import hash_password
from ..rag_engine import rag_engine
from ..ingest_knowledge_base import ingest_directory

# Define absolute path for knowledge base to ensure consistency
BACKEND_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
KB_DIR = os.path.join(BACKEND_DIR, "knowledge_base")

# Ensure directory exists
if not os.path.exists(KB_DIR):
    os.makedirs(KB_DIR)

router = APIRouter(
    prefix="/admin",
    tags=["admin"],
    responses={404: {"description": "Not found"}},
)

@router.get("/stats")
def get_system_stats(
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """Get system-wide statistics for the dashboard"""
    total_users = db.query(User).count()
    active_users = db.query(User).filter(User.is_active == True).count()
    admin_users = db.query(User).filter(User.role == "admin").count()
    
    # ChromaDB stats
    kb_stats = rag_engine.get_collection_stats()
    
    # Simple "health" check
    try:
        cpu_usage = psutil.cpu_percent()
        memory_usage = psutil.virtual_memory().percent
        disk_usage = psutil.disk_usage('.').percent
        status_msg = "Operational"
        if cpu_usage > 90 or memory_usage > 90:
            status_msg = "High Load"
    except:
        cpu_usage = 0
        memory_usage = 0
        disk_usage = 0
        status_msg = "Unknown"

    return {
        "users": {
            "total": total_users,
            "active": active_users,
            "admins": admin_users
        },
        "knowledge_base": {
            "total_documents": kb_stats.get("total_documents", 0)
        },
        "system": {
            "status": status_msg,
            "cpu_load": cpu_usage,       # Aligned with frontend expectation
            "memory_usage": memory_usage, # Aligned with frontend expectation
            "disk": disk_usage
        }
    }

@router.get("/progress")
def get_all_trainee_progress(
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """Get aggregated and detailed progress for all trainees"""
    # Fetch all trainees
    users = db.query(User).filter(User.role != "admin").all()
    user_ids = [str(u.id) for u in users]
    
    # Batch fetch all progress and scores to avoid N+1 problem
    all_progress = db.query(UserProgress).filter(UserProgress.user_id.in_(user_ids)).all()
    all_scores = db.query(QuizScore).filter(QuizScore.user_id.in_(user_ids)).all()
    
    # Organize data by user_id
    progress_map = {}
    for p in all_progress:
        if p.user_id not in progress_map:
            progress_map[p.user_id] = []
        progress_map[p.user_id].append(p)
        
    scores_map = {}
    for s in all_scores:
        if s.user_id not in scores_map:
            scores_map[s.user_id] = []
        scores_map[s.user_id].append(s)
    
    results = []
    for user in users:
        uid_str = str(user.id)
        user_progress = progress_map.get(uid_str, [])
        user_scores = scores_map.get(uid_str, [])
        
        lessons = [
            {
                "course_id": p.course_id,
                "percentage": p.progress_percentage,
                "last_accessed": p.last_accessed
            } for p in user_progress
        ]
        
        quizzes = [
            {
                "course_id": q.course_id,
                "lesson_id": q.lesson_id,
                "score": q.score,
                "first_attempt_score": q.first_attempt_score,
                "attempts_count": q.attempts_count,
                "completed_at": q.completed_at,
                "first_attempt_at": q.first_attempt_at
            } for q in user_scores
        ]
        
        # Calculate Weighted Mastery Index
        # Formula: (Best Score) * Efficiency Factor
        # Efficiency Factor: 1.0 (1-2 attempts), 0.9 (3-5), 0.75 (6-9), 0.6 (10+)
        weighted_scores = []
        for q in user_scores:
            efficiency_factor = 1.0
            if q.attempts_count > 9: efficiency_factor = 0.6
            elif q.attempts_count >= 6: efficiency_factor = 0.75
            elif q.attempts_count >= 3: efficiency_factor = 0.9
            
            weighted_scores.append(q.score * efficiency_factor)
            
        avg_score = 0
        if weighted_scores:
            avg_score = sum(weighted_scores) / len(weighted_scores)
        
        results.append({
            "id": user.id,
            "username": user.username,
            "full_name": user.full_name,
            "last_login": user.last_login,
            "completed_lessons": len(user_progress),
            "average_score": round(float(avg_score), 1), # This is now the Weighted Mastery Index
            "raw_average_score": round(float(sum(q.score for q in user_scores) / len(user_scores)), 1) if user_scores else 0,
            "lessons_history": lessons,
            "quizzes_history": quizzes
        })
        
    return results

@router.get("/logs", response_model=List[Dict])
def get_system_logs(
    limit: int = 50,
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """Get recent system audit logs"""
    logs = db.query(SystemLog).order_by(SystemLog.created_at.desc()).limit(limit).all()
    return [
        {
            "id": log.id,
            "level": log.level,
            "message": log.message,
            "context": log.context,
            "created_at": log.created_at
        } for log in logs
    ]

@router.delete("/users/{user_id}")
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """Permanently delete a user. Admin only."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if user.id == admin.id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
        
    # Log the deletion
    log_entry = SystemLog(
        level="WARNING",
        message=f"Admin {admin.username} deleted user {user.username}",
        context="USER_MGMT",
        user_id=admin.id
    )
    db.add(log_entry)
    
    db.delete(user)
    db.commit()
    
    return {"message": f"User {user_id} deleted successfully"}


@router.post("/users", response_model=UserResponse)
def create_user_as_admin(
    user_data: UserCreateAdmin,
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """Admin-only endpoint to create users with direct role assignment."""
    # Check for existing user
    if db.query(User).filter(User.username == user_data.username).first():
        raise HTTPException(status_code=400, detail="Username already registered")
    if db.query(User).filter(User.email == user_data.email).first():
        raise HTTPException(status_code=400, detail="Email already registered")

    new_user = User(
        username=user_data.username,
        email=user_data.email,
        hashed_password=hash_password(user_data.password),
        full_name=user_data.full_name,
        role=user_data.role,
        is_active=True
    )
    
    db.add(new_user)
    
    # Log creation
    log_entry = SystemLog(
        level="INFO",
        message=f"Admin {admin.username} created user {new_user.username} as {new_user.role}",
        context="USER_MGMT",
        user_id=admin.id
    )
    db.add(log_entry)
    db.commit()
    db.refresh(new_user)
    
    return new_user


@router.put("/users/{user_id}", response_model=UserResponse)
def update_user_as_admin(
    user_id: int,
    user_update: UserUpdate,
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """Admin-only endpoint to update user details."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Update basic fields
    if user_update.email is not None:
        user.email = user_update.email
    if user_update.full_name is not None:
        user.full_name = user_update.full_name
    if user_update.role is not None:
        # Restriction: admin cannot demote themselves to avoid locking out the system
        if user.id == admin.id and user_update.role != "admin":
             raise HTTPException(status_code=400, detail="Cannot demote yourself from Admin role")
        user.role = user_update.role
    if user_update.is_active is not None:
        if user.id == admin.id and user_update.is_active is False:
             raise HTTPException(status_code=400, detail="Cannot deactivate your own account")
        user.is_active = user_update.is_active
    
    # Handle password update separately (re-hash)
    if user_update.password:
        user.hashed_password = hash_password(user_update.password)

    # Log update
    log_entry = SystemLog(
        level="INFO",
        message=f"Admin {admin.username} updated user {user.username} details",
        context="USER_MGMT",
        user_id=admin.id
    )
    db.add(log_entry)
    db.commit()
    db.refresh(user)
    
    return user


@router.post("/reopen-assessment")
def reopen_assessment(
    user_id: int,
    quiz_slug: str,
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """
    Delete a trainee's quiz score and question attempts for a specific assessment.
    This effectively 'reopens' the assessment for them to take again.
    """
    # 1. Find the Quiz to get its numeric ID (needed for QuestionAttempt)
    quiz = db.query(Quiz).filter(Quiz.slug == quiz_slug).first()
    
    # 2. Delete the QuizScore (uses user_id as string and lesson_id as slug)
    score = db.query(QuizScore).filter(
        QuizScore.user_id == str(user_id),
        QuizScore.lesson_id == quiz_slug
    ).first()
    
    if score:
        db.delete(score)
        
    # 3. Delete QuestionAttempts (uses numeric IDs)
    if quiz:
        db.query(QuestionAttempt).filter(
            QuestionAttempt.user_id == user_id,
            QuestionAttempt.quiz_id == quiz.id
        ).delete()
    
    # 4. Log the action
    log_entry = SystemLog(
        level="WARNING",
        message=f"Admin {admin.username} reopened assessment '{quiz_slug}' for user ID {user_id}",
        context="ASSESSMENT_MGMT",
        user_id=admin.id
    )
    db.add(log_entry)
    
    db.commit()
    
    return {"message": f"Assessment '{quiz_slug}' has been reopened for user {user_id}"}


@router.post("/reopen-all-assessments")
def reopen_all_assessments(
    user_id: int,
    course_type: str = None,  # "2D_Drawing" or "3D_Modeling" or None for all
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """
    Delete quiz scores and question attempts for a trainee.
    Optionally filtered by course_type.
    """
    if course_type:
        # 1. Delete QuizScore entries for the specific course_id
        # Note: course_id is "1" for 3D and "2" for 2D in the models
        internal_course_id = "1" if course_type == "3D_Modeling" else "2"
        db.query(QuizScore).filter(
            QuizScore.user_id == str(user_id),
            QuizScore.course_id == internal_course_id
        ).delete()
        
        # 2. Delete QuestionAttempts for quizzes in that course
        quizzes = db.query(Quiz).filter(Quiz.course_type == course_type).all()
        quiz_ids = [q.id for q in quizzes]
        if quiz_ids:
            db.query(QuestionAttempt).filter(
                QuestionAttempt.user_id == user_id,
                QuestionAttempt.quiz_id.in_(quiz_ids)
            ).delete()
    else:
        # Delete all
        db.query(QuizScore).filter(QuizScore.user_id == str(user_id)).delete()
        db.query(QuestionAttempt).filter(QuestionAttempt.user_id == user_id).delete()
    
    # 3. Log the action
    msg = f"Admin {admin.username} reset {'ALL' if not course_type else course_type} assessments for user ID {user_id}"
    log_entry = SystemLog(
        level="DANGER",
        message=msg,
        context="ASSESSMENT_MGMT",
        user_id=admin.id
    )
    db.add(log_entry)
    
    db.commit()
    
    return {"message": f"Assessments have been reopened for user {user_id}"}


@router.post("/close-all-assessments")
def close_all_assessments(
    user_id: int,
    course_type: str = None, # "2D_Drawing" or "3D_Modeling" or None for all
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """
    Mark assessments as completed with 100% score for a trainee.
    Optionally filtered by course_type.
    """
    # 1. Get relevant quizzes
    query = db.query(Quiz)
    if course_type:
        query = query.filter(Quiz.course_type == course_type)
    quizzes = query.all()
    
    # 2. For each quiz, upsert a QuizScore entry
    for quiz in quizzes:
        # Check if score exists
        score_entry = db.query(QuizScore).filter(
            QuizScore.user_id == str(user_id),
            QuizScore.lesson_id == quiz.slug
        ).first()
        
        if score_entry:
            score_entry.score = 100.0
            score_entry.attempts_count = max(score_entry.attempts_count, 1)
            score_entry.completed_at = datetime.utcnow()
        else:
            new_score = QuizScore(
                user_id=str(user_id),
                course_id="1" if quiz.course_type == "3D_Modeling" else "2",
                lesson_id=quiz.slug,
                score=100.0,
                attempts_count=1,
                completed_at=datetime.utcnow()
            )
            db.add(new_score)
            
    # 3. Log the action
    msg = f"Admin {admin.username} marked {'ALL' if not course_type else course_type} assessments as CLOSED (100%) for user ID {user_id}"
    log_entry = SystemLog(
        level="INFO",
        message=msg,
        context="ASSESSMENT_MGMT",
        user_id=admin.id
    )
    db.add(log_entry)
    
    db.commit()
    
    return {"message": f"Assessments have been marked as completed for user {user_id}"}


@router.post("/broadcast")
def create_broadcast(
    message: str,
    level: str = "info",
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """Create a new system-wide broadcast message"""
    broadcast = Broadcast(message=message, level=level, created_by=admin.id)
    db.add(broadcast)
    db.commit()
    return {"message": "Broadcast sent successfully"}


@router.get("/broadcasts/active")
def get_active_broadcasts(
    db: Session = Depends(get_db)
):
    """Get all currently active broadcasts with sender name"""
    broadcasts = db.query(Broadcast, User.full_name)\
        .join(User, Broadcast.created_by == User.id)\
        .filter(Broadcast.is_active == True)\
        .order_by(Broadcast.created_at.desc()).all()
    
    return [
        {
            "id": b[0].id,
            "message": b[0].message,
            "level": b[0].level,
            "created_at": b[0].created_at,
            "sender_name": b[1]
        } for b in broadcasts
    ]


@router.get("/heatmap")
def get_training_heatmap(
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """Get activity counts per course for heatmap (Recent 60 mins)"""
    from datetime import datetime, timedelta
    one_hour_ago = datetime.utcnow() - timedelta(minutes=60)
    
    stats = db.query(
        UserProgress.course_id,
        func.count(UserProgress.id).label("active_count")
    ).filter(UserProgress.last_accessed >= one_hour_ago)\
     .group_by(UserProgress.course_id).all()
    
    return [{"course_id": s[0], "count": s[1]} for s in stats]


@router.get("/export/progress")
def export_trainee_progress(
    user_id: int = None,
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """Export granular trainee progress data as CSV"""
    import csv
    import io
    from fastapi.responses import StreamingResponse
    
    # Optimized query: Fetch progress and join with users
    query = db.query(UserProgress, User)\
        .join(User, UserProgress.user_id == cast(User.id, String))
        
    if user_id:
        query = query.filter(User.id == user_id)
        
    results = query.all()
    
    if not results and user_id:
        # If no progress entries, at least check if user exists
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        # Generate results for user with no progress
        results = [(None, user)]
    elif not results:
        raise HTTPException(status_code=404, detail="No progress data found to export")
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Trainee Username", "Full Name", "Course ID", "Status", 
        "Progress %", "Highest Quiz Score", "Last Activity"
    ])
    
    for progress, user in results:
        username = user.username
        full_name = user.full_name
        course_id = progress.course_id if progress else "N/A"
        progress_pct = progress.progress_percentage if progress else 0
        last_act = progress.last_accessed if progress else "N/A"
        
        # Get highest quiz score for this user-course pair
        best_score = 0
        if progress:
            best_score = db.query(func.max(QuizScore.score))\
                .filter(QuizScore.user_id == str(user.id))\
                .filter(QuizScore.course_id == progress.course_id).scalar() or 0
            
        status = "Completed" if progress_pct >= 100 else "In Progress"
        
        writer.writerow([
            username,
            full_name,
            course_id,
            status,
            f"{progress_pct}%",
            f"{best_score}%",
            last_act
        ])
    
    output.seek(0)
    filename = f"trainee_granular_report_{user_id if user_id else 'all'}.csv"
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/reindex")
def trigger_kb_reindex(
    admin: User = Depends(require_role("admin"))
):
    """Manually trigger re-indexing of the knowledge base directory"""
    kb_dir = KB_DIR
    try:
        ingest_directory(kb_dir)
        return {"message": "Knowledge base re-indexing triggered successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Re-indexing failed: {str(e)}")

@router.get("/kb/files")
def list_kb_files(
    admin: User = Depends(require_role("admin"))
):
    """List all files in the knowledge base directory"""
    kb_dir = KB_DIR
    files = []
    for filename in os.listdir(kb_dir):
        file_path = os.path.join(kb_dir, filename)
        if os.path.isfile(file_path):
            stats = os.stat(file_path)
            files.append({
                "name": filename,
                "size": stats.st_size,
                "modified": datetime.fromtimestamp(stats.st_mtime)
            })
    return files

@router.post("/kb/upload")
async def upload_kb_files(
    files: List[UploadFile] = File(...),
    admin: User = Depends(require_role("admin"))
):
    """Upload one or more files to the knowledge base"""
    kb_dir = KB_DIR
    uploaded_files = []
    for file in files:
        # Secure filename or at least check extension
        if not file.filename.endswith(('.xlsx', '.csv', '.py')):
            continue
            
        file_path = os.path.join(kb_dir, file.filename)
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        uploaded_files.append(file.filename)
        
    return {"message": f"Uploaded {len(uploaded_files)} files: {', '.join(uploaded_files)}"}

@router.get("/kb/files/{filename}/preview")
def preview_kb_file(
    filename: str,
    admin: User = Depends(require_role("employee"))
):
    """Return file contents as JSON rows for in-browser preview (CSV and XLSX only)."""
    # Sanitize — no path traversal
    if "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")

    file_path = os.path.join(KB_DIR, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")

    ext = filename.lower().rsplit(".", 1)[-1]

    try:
        if ext == "csv":
            rows = []
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames or []
                for row in reader:
                    rows.append(dict(row))
            return {"filename": filename, "type": "csv", "headers": list(headers), "rows": rows[:500]}

        elif ext == "xlsx":
            try:
                import openpyxl
            except ImportError:
                raise HTTPException(status_code=500, detail="openpyxl not installed")
            
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheets = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_data = []
                headers = []
                # Determine max columns for headers
                max_cols = ws.max_column or 0
                
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        # Normalize headers
                        headers = [str(c) if c is not None and str(c).strip() != "" else f"Column {idx+1}" for idx, c in enumerate(row)]
                    else:
                        # Safely map values to available headers
                        row_dict = {}
                        for j, value in enumerate(row):
                            if j < len(headers):
                                row_dict[headers[j]] = str(value) if value is not None else ""
                            else:
                                # New column not defined in header, skip or add to dynamic headers? 
                                # Let's skip to keep table consistent
                                pass
                        rows_data.append(row_dict)
                    
                    if i >= 500: # Increase limit slightly or keep at 500
                        break
                
                sheets[sheet_name] = {"headers": headers, "rows": rows_data}
            wb.close()
            return {"filename": filename, "type": "xlsx", "sheets": sheets}

        else:
            raise HTTPException(status_code=400, detail="Only CSV and XLSX preview supported")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Preview failed: {str(e)}")


@router.get("/kb/files/{filename}/download")
def download_kb_file(
    filename: str,
    admin: User = Depends(require_role("employee"))
):
    """Download a KB file directly."""
    if "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    file_path = os.path.join(KB_DIR, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(
        path=file_path,
        filename=filename,
        media_type="application/octet-stream"
    )


@router.delete("/kb/files/{filename}")
def delete_kb_file(
    filename: str,
    admin: User = Depends(require_role("admin"))
):
    """Delete a file from the knowledge base"""
    file_path = os.path.join(KB_DIR, filename)
    
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")
        
    try:
        os.remove(file_path)
        return {"message": f"File {filename} deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to delete file: {str(e)}")


# ── Chat Log Monitoring ──────────────────────────────────────────────────────

@router.get("/chat-logs")
def get_chat_logs(
    limit: int = 50,
    offset: int = 0,
    username: str = None,
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """Paginated chat log list, optionally filtered by username."""
    query = db.query(ChatLog)
    if username:
        query = query.filter(ChatLog.username.ilike(f"%{username}%"))
    total = query.count()
    logs = query.order_by(ChatLog.created_at.desc()).offset(offset).limit(limit).all()
    return {
        "total": total,
        "offset": offset,
        "limit": limit,
        "logs": [
            {
                "id": l.id,
                "user_id": l.user_id,
                "username": l.username,
                "session_id": l.session_id,
                "message": l.message,
                "answer": l.answer,
                "sources_used": l.sources_used,
                "source_count": l.source_count,
                "tokens_estimated": l.tokens_estimated,
                "response_time_ms": l.response_time_ms,
                "had_media": l.had_media,
                "created_at": l.created_at.isoformat() if l.created_at else None,
            }
            for l in logs
        ]
    }


@router.get("/chat-logs/stats")
def get_chat_log_stats(
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """Aggregate stats for the Intelligence dashboard."""
    from sqlalchemy import func as sqlfunc

    total_queries = db.query(ChatLog).count()

    # Top 10 most active users
    top_users = (
        db.query(ChatLog.username, sqlfunc.count(ChatLog.id).label("count"))
        .group_by(ChatLog.username)
        .order_by(sqlfunc.count(ChatLog.id).desc())
        .limit(10)
        .all()
    )

    # Total estimated tokens used
    total_tokens = db.query(sqlfunc.sum(ChatLog.tokens_estimated)).scalar() or 0

    # Average response time
    avg_response_ms = db.query(sqlfunc.avg(ChatLog.response_time_ms)).scalar() or 0

    # Queries per day (last 14 days)
    from datetime import timedelta
    from sqlalchemy import cast, Date as SqlDate
    fourteen_days_ago = datetime.utcnow() - timedelta(days=14)
    daily = (
        db.query(
            cast(ChatLog.created_at, SqlDate).label("day"),
            sqlfunc.count(ChatLog.id).label("count")
        )
        .filter(ChatLog.created_at >= fourteen_days_ago)
        .group_by(cast(ChatLog.created_at, SqlDate))
        .order_by(cast(ChatLog.created_at, SqlDate))
        .all()
    )

    # Top 10 most queried sources
    source_counts: dict = {}
    all_sources = db.query(ChatLog.sources_used).filter(ChatLog.sources_used != None).all()
    for row in all_sources:
        for src in row[0].split(","):
            src = src.strip()
            if src:
                source_counts[src] = source_counts.get(src, 0) + 1
    top_sources = sorted(source_counts.items(), key=lambda x: x[1], reverse=True)[:10]

    return {
        "total_queries": total_queries,
        "total_tokens_estimated": int(total_tokens),
        "avg_response_ms": int(avg_response_ms),
        "top_users": [{"username": u, "count": c} for u, c in top_users],
        "queries_per_day": [{"day": str(d), "count": c} for d, c in daily],
        "top_sources": [{"source": s, "count": c} for s, c in top_sources],
    }


@router.delete("/chat-logs/{log_id}")
def delete_chat_log(
    log_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """Delete a single chat log entry."""
    log = db.query(ChatLog).filter(ChatLog.id == log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="Log not found")
    db.delete(log)
    db.commit()
    return {"message": "Deleted"}


# ── Feedback Endpoints ───────────────────────────────────────────────────

@router.get("/feedback/stats")
def get_feedback_stats(
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """Aggregate feedback stats: up/down totals, worst-rated queries."""
    from sqlalchemy import func as sqlfunc

    total_up = db.query(ChatFeedback).filter(ChatFeedback.rating == "up").count()
    total_down = db.query(ChatFeedback).filter(ChatFeedback.rating == "down").count()

    # Worst rated: log entries with most thumbs-down
    worst = (
        db.query(
            ChatFeedback.chat_log_id,
            sqlfunc.count(ChatFeedback.id).label("down_count")
        )
        .filter(ChatFeedback.rating == "down")
        .group_by(ChatFeedback.chat_log_id)
        .order_by(sqlfunc.count(ChatFeedback.id).desc())
        .limit(10)
        .all()
    )

    worst_with_messages = []
    for log_id, down_count in worst:
        log = db.query(ChatLog).filter(ChatLog.id == log_id).first()
        worst_with_messages.append({
            "log_id": log_id,
            "message": log.message if log else "(deleted)",
            "down_count": down_count
        })

    return {
        "total_up": total_up,
        "total_down": total_down,
        "satisfaction_rate": round(total_up / (total_up + total_down) * 100, 1) if (total_up + total_down) > 0 else None,
        "worst_rated": worst_with_messages
    }


# ── Cache Endpoints ─────────────────────────────────────────────────────

@router.get("/cache/stats")
def get_cache_stats(
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """Cache stats: total entries, total hits saved, top cached queries."""
    from sqlalchemy import func as sqlfunc

    total_entries = db.query(QueryCache).count()
    active_entries = db.query(QueryCache).filter(QueryCache.expires_at > datetime.utcnow()).count()
    total_hits = db.query(sqlfunc.sum(QueryCache.hit_count)).scalar() or 0

    top_cached = (
        db.query(QueryCache)
        .filter(QueryCache.expires_at > datetime.utcnow())
        .order_by(QueryCache.hit_count.desc())
        .limit(10)
        .all()
    )

    return {
        "total_entries": total_entries,
        "active_entries": active_entries,
        "total_hits_saved": int(total_hits),
        "top_cached_queries": [
            {
                "query": e.query_text,
                "hits": e.hit_count,
                "expires_at": e.expires_at.isoformat() if e.expires_at else None
            }
            for e in top_cached
        ]
    }


@router.delete("/cache")
def clear_cache(
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """Clear all cache entries (force-refresh everything)."""
    deleted = db.query(QueryCache).delete()
    db.commit()
    return {"message": f"Cleared {deleted} cache entries"}


# ── Assessment Management ───────────────────────────────────────────────────

@router.get("/quizzes", response_model=List[QuizResponse])
def get_all_quizzes(
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """List all quizzes with their metadata"""
    return db.query(Quiz).all()

@router.get("/quizzes/{quiz_id}", response_model=QuizResponse)
def get_quiz_detail(
    quiz_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """Get full quiz details including all questions"""
    quiz = db.query(Quiz).filter(Quiz.id == quiz_id).first()
    if not quiz:
        raise HTTPException(status_code=404, detail="Quiz not found")
    
    # Manually attach questions sorted by order
    questions = db.query(Question).filter(Question.quiz_id == quiz_id).order_by(Question.order).all()
    quiz.questions = questions
    return quiz

@router.post("/quizzes", response_model=QuizResponse)
def create_quiz(
    quiz_data: QuizCreate,
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """Create a new quiz definition"""
    if db.query(Quiz).filter(Quiz.slug == quiz_data.slug).first():
        raise HTTPException(status_code=400, detail="Quiz slug already exists")
    
    new_quiz = Quiz(**quiz_data.model_dump())
    db.add(new_quiz)
    db.commit()
    db.refresh(new_quiz)
    return new_quiz

@router.put("/quizzes/{quiz_id}", response_model=QuizResponse)
def update_quiz(
    quiz_id: int,
    quiz_update: QuizUpdate,
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """Update quiz metadata"""
    quiz = db.query(Quiz).filter(Quiz.id == quiz_id).first()
    if not quiz:
        raise HTTPException(status_code=404, detail="Quiz not found")
    
    for key, value in quiz_update.model_dump(exclude_unset=True).items():
        setattr(quiz, key, value)
    
    db.commit()
    db.refresh(quiz)
    return quiz


@router.delete("/quizzes/{quiz_id}")
def delete_quiz(
    quiz_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """Delete a quiz and all its questions"""
    quiz = db.query(Quiz).filter(Quiz.id == quiz_id).first()
    if not quiz:
        raise HTTPException(status_code=404, detail="Quiz not found")
    
    db.delete(quiz)
    db.commit()
    return {"message": "Quiz deleted successfully"}

# Question Endpoints
@router.post("/quizzes/{quiz_id}/questions", response_model=QuestionResponse)
def create_question(
    quiz_id: int,
    question_data: QuestionCreate,
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """Add a question to a quiz"""
    quiz = db.query(Quiz).filter(Quiz.id == quiz_id).first()
    if not quiz:
        raise HTTPException(status_code=404, detail="Quiz not found")
    
    new_question = Question(quiz_id=quiz_id, **question_data.model_dump())
    db.add(new_question)
    db.commit()
    db.refresh(new_question)
    return new_question

@router.put("/questions/{question_id}", response_model=QuestionResponse)
def update_question(
    question_id: int,
    question_update: QuestionUpdate,
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """Update an existing question"""
    question = db.query(Question).filter(Question.id == question_id).first()
    if not question:
        raise HTTPException(status_code=404, detail="Question not found")
    
    for key, value in question_update.model_dump(exclude_unset=True).items():
        setattr(question, key, value)
    
    db.commit()
    db.refresh(question)
    return question

@router.delete("/questions/{question_id}")
def delete_question(
    question_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """Remove a question from a quiz"""
    question = db.query(Question).filter(Question.id == question_id).first()
    if not question:
        raise HTTPException(status_code=404, detail="Question not found")
    
    db.delete(question)
    db.commit()
    return {"message": "Question deleted successfully"}


# ── Curriculum Management ───────────────────────────────────────────────────

@router.get("/curriculum/courses", response_model=List[CourseResponse])
def admin_get_courses(db: Session = Depends(get_db), admin: User = Depends(require_role("admin"))):
    return db.query(Course).order_by(Course.order).all()

@router.get("/curriculum/courses/{course_id}/lessons", response_model=List[LessonResponse])
def admin_get_lessons(course_id: int, db: Session = Depends(get_db), admin: User = Depends(require_role("admin"))):
    return db.query(Lesson).filter(Lesson.course_id == course_id).order_by(Lesson.order).all()

@router.post("/curriculum/lessons", response_model=LessonResponse)
def create_lesson(lesson_data: LessonCreate, db: Session = Depends(get_db), admin: User = Depends(require_role("admin"))):
    new_lesson = Lesson(**lesson_data.model_dump())
    db.add(new_lesson)
    db.commit()
    db.refresh(new_lesson)
    return new_lesson

@router.get("/curriculum/lessons/{lesson_id}/content", response_model=List[LessonContentResponse])
def get_lesson_content(lesson_id: int, db: Session = Depends(get_db), admin: User = Depends(require_role("admin"))):
    return db.query(LessonContent).filter(LessonContent.lesson_id == lesson_id).order_by(LessonContent.order).all()

@router.post("/curriculum/lessons/{lesson_id}/content", response_model=LessonContentResponse)
def add_lesson_content(lesson_id: int, content_data: LessonContentCreate, db: Session = Depends(get_db), admin: User = Depends(require_role("admin"))):
    new_content = LessonContent(**content_data.model_dump())
    db.add(new_content)
    db.commit()
    db.refresh(new_content)
    return new_content

# ── Analytics ─────────────────────────────────────────────────────────────

@router.get("/analytics/question-performance")
def get_question_performance(db: Session = Depends(get_db), admin: User = Depends(require_role("admin"))):
    """Analyze which questions trainees are failing most often"""
    # Use sqlalchemy func and cast
    from sqlalchemy import cast, Integer
    
    perf = db.query(
        Question.text,
        Quiz.title.label("quiz_title"),
        func.count(QuestionAttempt.id).label("total_attempts"),
        func.sum(cast(QuestionAttempt.is_correct, Integer)).label("correct_count")
    ).join(QuestionAttempt, Question.id == QuestionAttempt.question_id)\
     .join(Quiz, Quiz.id == Question.quiz_id)\
     .group_by(Question.id, Quiz.id).all()
     
    results = []
    for p in perf:
        accuracy = (p.correct_count / p.total_attempts * 100) if p.total_attempts > 0 else 0
        results.append({
            "question": p.text,
            "quiz": p.quiz_title,
            "attempts": p.total_attempts,
            "accuracy": round(float(accuracy), 1),
            "is_critical": accuracy < 60 and p.total_attempts > 5
        })
    
    return sorted(results, key=lambda x: x["accuracy"])


@router.get("/trainee/{user_id}/attempts/{quiz_slug}")
def get_trainee_quiz_attempts(
    user_id: int,
    quiz_slug: str,
    db: Session = Depends(get_db),
    admin: User = Depends(require_role("admin"))
):
    """Get detailed question-by-question attempts for a trainee in a specific quiz"""
    # 1. Find the quiz
    quiz = db.query(Quiz).filter(Quiz.slug == quiz_slug).first()
    if not quiz:
        raise HTTPException(status_code=404, detail="Quiz not found")
        
    # 2. Get questions for this quiz
    questions = db.query(Question).filter(Question.quiz_id == quiz.id).order_by(Question.order).all()
    
    # 3. Get all attempts for this user and quiz
    attempts = db.query(QuestionAttempt).filter(
        QuestionAttempt.user_id == user_id,
        QuestionAttempt.quiz_id == quiz.id
    ).order_by(QuestionAttempt.attempted_at.desc()).all()
    
    # Map latest attempts to question IDs
    attempts_map = {}
    for a in attempts:
        if a.question_id not in attempts_map:
            attempts_map[a.question_id] = a
    
    results = []
    for q in questions:
        attempt = attempts_map.get(q.id)
        
        # Parse options safely
        try:
            options = json.loads(q.options_json) if q.options_json else []
        except:
            options = []
            
        results.append({
            "question_text": q.text,
            "explanation": q.explanation,
            "options": options,
            "correct_answer_index": q.correct_answer,
            "user_answer_index": attempt.chosen_option if attempt else None,
            "is_correct": attempt.is_correct if attempt else False,
            "seconds_spent": attempt.seconds_spent if attempt else 0,
            "attempted_at": attempt.attempted_at if attempt else None
        })
        
    return {
        "quiz_title": quiz.title,
        "attempts": results
    }

